import os
import re
import json
import math
import email
import sqlite3
import logging
import traceback
import urllib.request
import urllib.error
from dotenv import load_dotenv
import requests
from email import policy
import uuid
import random
import string
import hmac
import hashlib
import time
import base64
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor

from flask import Flask, request, jsonify
from bs4 import BeautifulSoup

project_folder = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(project_folder, '.env'))


# ==============================================================================
#  RATE LIMITER (in-memory, per-IP)
#  Sliding window: max 20 predict calls per 60 seconds per IP.
#  NOTE: in-memory store is per-worker on multi-worker hosts, so the effective
#  limit may be higher than RATE_LIMIT_MAX under load.
# ==============================================================================
_rate_store = defaultdict(list)   # ip -> [timestamp, ...]
RATE_LIMIT_MAX  = 20               # max requests
RATE_LIMIT_SECS = 60              # per window

def _check_rate_limit(ip: str) -> bool:
    """Returns True if request is allowed, False if rate-limited."""
    now = time.time()
    _rate_store[ip] = [t for t in _rate_store[ip] if now - t < RATE_LIMIT_SECS]
    if len(_rate_store[ip]) >= RATE_LIMIT_MAX:
        return False
    _rate_store[ip].append(now)
    return True

  # Read-Only: 28k rows of cutoffs & colleges
USER_DB_PATH = 'Users.db'         # Read/Write: SaaS Accounts, Tokens, Economy
# ------------------------------------------------------------------ logging ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("score-engine")

# --------------------------------------------------- optional dependencies ----
# The core scoring engine MUST work even if google-genai is not
# installed. They are imported lazily so a missing package never kills the app.
try:
    from flask_cors import CORS
    _HAS_CORS = True
except Exception:                                       # pragma: no cover
    _HAS_CORS = False
    log.warning("flask_cors not installed - CORS headers added manually.")

try:
    from google import genai
    from google.genai import types
    _HAS_GENAI = True
except Exception:                                       # pragma: no cover
    _HAS_GENAI = False
    log.warning("google-genai not installed - using rule-based advice fallback.")


# ==============================================================================
#  APP + CONFIG
# ==============================================================================
app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 64 * 1024 * 1024     # 64 MB upload ceiling

if _HAS_CORS:
    CORS(app)
else:
    @app.after_request
    def _add_cors(resp):
        resp.headers["Access-Control-Allow-Origin"] = "*"
        resp.headers["Access-Control-Allow-Headers"] = "*"
        resp.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
        return resp

# --- API keys ---------------------------------------------------------------
# ┌─────────────────────────────────────────────────────────────────────────┐
# │  GEMINI KEY — stored as an environment variable / .env file.            │
# │  Get it free at: https://aistudio.google.com/app/apikey                │
# └─────────────────────────────────────────────────────────────────────────┘

# Load the hidden keys from the .env file
load_dotenv()

# Grab the Gemini key from the environment. The actual client is created
# below in the `_ai_client = genai.Client(...)` block.
GEMINI_KEY = os.getenv("GEMINI_API_KEY")

# --- CASHFREE CONFIGURATION -------------------------------------------------
# Get these from your Cashfree dashboard → Developers → API Keys.
# You can also set them as environment variables CASHFREE_APP_ID / CASHFREE_SECRET_KEY.
CASHFREE_APP_ID  = os.environ.get("CASHFREE_APP_ID")
CASHFREE_SECRET_KEY = os.environ.get("CASHFREE_SECRET_KEY")

# ┌─────────────────────────────────────────────────────────────────────────┐
# │  ENVIRONMENT SWITCH — must match `mode` in your HTML files!              │
# │  "production" = real money.  "sandbox" = test cards.                     │
# │  Whatever you set here, set the SAME value in premium.html, predict.html │
# │  and booking.html → Cashfree({ mode: "..." }).                           │
# └─────────────────────────────────────────────────────────────────────────┘
CASHFREE_ENV = os.environ.get("CASHFREE_ENV")

# Cashfree uses different hostnames for sandbox vs production.
_CF_HOST = "https://sandbox.cashfree.com" if CASHFREE_ENV == "sandbox" else "https://api.cashfree.com"
CASHFREE_URL = f"{_CF_HOST}/pg/orders"

CASHFREE_HEADERS = {
    "accept": "application/json",
    "x-client-id": CASHFREE_APP_ID,
    "x-client-secret": CASHFREE_SECRET_KEY,
    "x-api-version": "2023-08-01",
    "content-type": "application/json",
}


# True only when real keys have been pasted in (controls the 503 "disabled" guard).
_HAS_CASHFREE = bool(
    CASHFREE_APP_ID and CASHFREE_SECRET_KEY
    and "YOUR_CASHFREE" not in CASHFREE_APP_ID
    and "YOUR_CASHFREE" not in CASHFREE_SECRET_KEY
)
if not _HAS_CASHFREE:
    log.warning("Cashfree keys not set - payment endpoints will return 503.")

_ai_client = None
if _HAS_GENAI and GEMINI_KEY:
    try:
        _ai_client = genai.Client(api_key=GEMINI_KEY)
        log.info("✅ Gemini client ready.  Advice=gemini-2.5-flash  Placement=gemini-2.5-pro")
    except Exception as e:
        log.warning("Gemini init failed (%s) - rule-based advice fallback.", e)
else:
    if not _HAS_GENAI:
        log.warning("⚠️  google-genai not installed.  Run: pip install google-genai")
    else:
        log.warning("⚠️  No GEMINI_API_KEY found — using rule-based fallback."
                    "  Open backend.py and paste your key into GEMINI_KEY_HARDCODED.")

# ==============================================================================
#  CASHFREE HELPERS
# ==============================================================================
def _cf_create_order(amount_rupees, email, order_note="gmc_pro_order"):
    """
    Creates a Cashfree order and returns (payment_session_id, order_id).
    Raises Exception on failure.
    """
    order_id = f"GMC_{uuid.uuid4().hex[:10].upper()}"
    safe_email = (email or "student@example.com").strip()
    customer_id = (safe_email.split('@')[0][:40]) or "guest"

    payload = {
        "order_id": order_id,
        "order_amount": float(amount_rupees),
        "order_currency": "INR",
        "order_note": order_note,
        "customer_details": {
            "customer_id": customer_id,
            "customer_phone": "9999999999",   # Cashfree requires a phone number
            "customer_email": safe_email,
        },
        "order_meta": {
                # Cashfree calls this URL server-to-server when payment completes,
                # even if the user closes their browser. Set it to your live domain.
                "notify_url": "https://parthsancheti.pythonanywhere.com/cashfree-webhook"
            }
        }

    resp = requests.post(CASHFREE_URL, json=payload, headers=CASHFREE_HEADERS, timeout=30)
    data = resp.json()
    if resp.status_code == 200 and "payment_session_id" in data:
        return data["payment_session_id"], order_id
    raise Exception(f"Cashfree order failed [{resp.status_code}]: {data}")


def _cf_order_is_paid(order_id):
    """Asks Cashfree for the real status of an order. Returns True if PAID."""
    verify_url = f"{CASHFREE_URL}/{order_id}"
    # GET requests don't send a body, so drop content-type.
    verify_headers = {k: v for k, v in CASHFREE_HEADERS.items() if k != "content-type"}
    resp = requests.get(verify_url, headers=verify_headers, timeout=30)
    data = resp.json()
    return resp.status_code == 200 and data.get("order_status") == "PAID"


# ==============================================================================
#  EXAM CONFIGURATION
# ==============================================================================
# Per-exam marking rules. `weights` maps a (normalised) subject -> marks/question.
# `default_weight` covers any subject not explicitly listed.
EXAM_CONFIG = {
    "MHT-CET": {
        "label": "MHT-CET",
        "negative": 0.0,                       # MHT-CET has NO negative marking
        "default_weight": 1,
        "weights": {"Mathematics": 2},         # Maths questions are worth 2
        "scheme": "MHT-CET: +1 per correct (Physics/Chemistry/Biology), "
                  "+2 per correct (Mathematics), no negative marking.",
    },
    "NEET": {
        "label": "NEET (UG)",
        "negative": 1.0,                       # NEET: -1 for a wrong answer
        "default_weight": 4,                   # +4 for a correct answer
        "weights": {},
        "scheme": "NEET: +4 per correct, -1 per incorrect, 0 if unanswered.",
    },
}

# Canonical subject names so PHYSICS / Physics / physics all collapse to one key.
_SUBJECT_ALIASES = {
    "physics": "Physics",
    "chemistry": "Chemistry",
    "maths": "Mathematics", "math": "Mathematics", "mathematics": "Mathematics",
    "biology": "Biology", "bio": "Biology",
    "botany": "Botany",
    "zoology": "Zoology",
}


def normalise_subject(raw):
    """'PHYSICS ' -> 'Physics'. Unknown labels are title-cased and kept as-is."""
    if not raw:
        return "General"
    key = re.sub(r"[^a-z]", "", raw.strip().lower())
    return _SUBJECT_ALIASES.get(key, raw.strip().title() or "General")


# ==============================================================================
#  CUSTOM ERROR TYPE  ->  always produces a clean JSON body
# ==============================================================================
class ParseError(Exception):
    """Raised for any expected, user-facing failure (bad file, etc.)."""
    def __init__(self, code, message, detail="", status=400):
        super().__init__(message)
        self.code = code
        self.message = message
        self.detail = detail
        self.status = status


def error_payload(code, message, detail="", status=400):
    body = {
        "result": "error",
        "error_code": code,
        "error": message,          # short, friendly - frontend shows this
        "detail": detail,          # technical context for the dev console
    }
    return jsonify(body), status


# ==============================================================================
#  STEP 1  -  TURN AN UPLOAD INTO CLEAN HTML
# ==============================================================================
def extract_html(raw_bytes, filename=""):
    """
    Accepts the raw bytes of an upload and returns a usable HTML string.

    Handles:
      * .mht / .mhtml  -> MIME multipart archive; we pull out the text/html part.
      * .html / .htm   -> decoded directly.
    Detection is content-based, so a mislabelled extension still works.
    """
    if not raw_bytes:
        raise ParseError("EMPTY_FILE", "The uploaded file is empty.",
                          "0 bytes received.")

    head = raw_bytes[:600].lstrip()
    looks_like_mht = (
        filename.lower().endswith((".mht", ".mhtml"))
        or head.startswith(b"From:")
        or b"multipart/related" in raw_bytes[:2000]
        or b"Snapshot-Content-Location" in raw_bytes[:2000]
    )

    if looks_like_mht:
        try:
            msg = email.message_from_bytes(raw_bytes, policy=policy.default)
        except Exception as e:
            raise ParseError("MHT_PARSE_FAILED",
                              "We could not read this .mht archive.",
                              f"email module error: {e}")
        html_parts = []
        for part in msg.walk():
            if part.get_content_type() == "text/html":
                try:
                    payload = part.get_payload(decode=True) or b""
                    html_parts.append(payload.decode("utf-8", errors="ignore"))
                except Exception:
                    continue
        if not html_parts:
            raise ParseError("MHT_NO_HTML",
                              "This .mht file contains no readable web page.",
                              "No text/html part inside the MIME archive.")
        # Prefer the part that actually contains question markup; else the biggest.
        for h in html_parts:
            if "tblObjection" in h or "menu-tbl" in h:
                return h
        return max(html_parts, key=len)

    # Plain HTML - try a couple of encodings before giving up.
    for enc in ("utf-8", "utf-16", "latin-1"):
        try:
            return raw_bytes.decode(enc)
        except Exception:
            continue
    return raw_bytes.decode("utf-8", errors="ignore")


def make_soup(html):
    """Parse with lxml when available, fall back to the stdlib parser."""
    try:
        return BeautifulSoup(html, "lxml")
    except Exception:
        return BeautifulSoup(html, "html.parser")


# ==============================================================================
#  STEP 2  -  DETECT WHICH EXAM THIS SHEET BELONGS TO
# ==============================================================================
def detect_exam_type(soup):
    """Return 'MHT-CET', 'NEET' or None based on structural fingerprints."""
    if soup.find("table", id="tblObjection"):
        return "MHT-CET"

    text = soup.get_text(" ", strip=True).lower()

    # MHT-CET objection sheet fingerprints.
    if "candidate response" in text and "correct option" in text:
        return "MHT-CET"
    if "mhexam" in text or "mht-cet" in text or "mht cet" in text:
        return "MHT-CET"

    # NTA / NEET fingerprints.
    if soup.find("table", class_=re.compile(r"menu-tbl", re.I)):
        return "NEET"
    if "chosen option" in text and "question id" in text:
        return "NEET"
    if "national testing agency" in text or "neet" in text:
        return "NEET"
    return None


# ==============================================================================
#  STEP 3a  -  MHT-CET PARSER
# ==============================================================================
def parse_mhtcet(soup):
    """
    Parse an MHT-CET Objection-Tracker response sheet.

    The sheet structure:
        <table id="tblObjection"><tbody>
            <tr> ...header... </tr>
            <tr><td>QID</td><td>SECTION</td><td> ...question...
                    <div class="BoxNumber">opt1 id</div> x4
                    <table class="...center">
                        <span>Correct Option id</span>
                        <span>Candidate Response id</span>
                    </table>
            </td><td>Raise Objection</td></tr>
            ...
        </tbody></table>

    Because the correct option lives inside the sheet itself, scoring is exact
    and works for ANY shift / year with no answer-key file.
    """
    table = soup.find("table", id="tblObjection")
    rows = []
    if table:
        body = table.find("tbody") or table
        rows = body.find_all("tr", recursive=False)
    if not rows:
        # Fallback: any table whose header row mentions Question ID + Section.
        for t in soup.find_all("table"):
            head = t.get_text(" ", strip=True).lower()
            if "question id" in head and "section" in head:
                body = t.find("tbody") or t
                rows = body.find_all("tr", recursive=False)
                if rows:
                    break
    if not rows:
        raise ParseError(
            "MHTCET_NO_QUESTIONS",
            "We could not find any MHT-CET questions in this file.",
            "No <table id='tblObjection'> or equivalent question table found.")

    questions = []
    UNANSWERED = {"", "-", "--", "---", "na", "n/a", "not answered",
                  "not attempted", "none"}

    for row in rows:
        cells = row.find_all("td", recursive=False)
        if len(cells) < 3:                  # header row / spacer -> skip
            continue

        q_id = cells[0].get_text(strip=True)
        section = normalise_subject(cells[1].get_text(strip=True))
        body_cell = cells[2]

        # The four option IDs, in display order (option 1..4).
        option_ids = [d.get_text(strip=True)
                      for d in body_cell.find_all("div", class_="BoxNumber")]

        # The inner result table holds Correct Option + Candidate Response.
        inner = body_cell.find("table", class_=re.compile(r"center", re.I))
        spans = inner.find_all("span") if inner else []
        if len(spans) < 2:
            # Not a real question row (could be a layout artefact) - skip safely.
            continue

        correct_raw = spans[0].get_text(strip=True)
        chosen_raw = spans[1].get_text(strip=True)

        # Map an option ID back to a human-friendly 1-4 index when possible.
        def to_index(val):
            if val in option_ids:
                return option_ids.index(val) + 1
            return None

        cancelled = correct_raw.lower() in UNANSWERED or "cancel" in correct_raw.lower()
        attempted = chosen_raw.lower() not in UNANSWERED

        if cancelled:
            status = "cancelled"      # grace marks - treated as correct below
        elif not attempted:
            status = "unanswered"
        elif chosen_raw == correct_raw:
            status = "correct"
        else:
            status = "incorrect"

        questions.append({
            "question_id": q_id or f"Q{len(questions) + 1}",
            "subject": section,
            "chosen_index": to_index(chosen_raw),
            "correct_index": to_index(correct_raw),
            "status": status,
        })

    if not questions:
        raise ParseError(
            "MHTCET_EMPTY",
            "No valid MHT-CET questions could be read from this sheet.",
            "Question table was present but contained 0 parseable rows.")
    return questions


# ==============================================================================
#  STEP 3b  -  NEET (NTA) PARSER
# ==============================================================================
def load_neet_answer_key():
    """
    NEET response sheets from the NTA do NOT contain the correct answer, so an
    external key is required for exact scoring. Drop a file named
    `neet_answer_key.json` next to this script:

        { "questionId": "correctOptionId",  ...  }      # option ID, OR
        { "questionId": 3, ... }                        # 1-based option number

    If the file is absent we still parse the sheet and report attempt counts,
    but total_score is returned as null with a clear warning.
    """
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "neet_answer_key.json")
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {str(k): str(v).strip() for k, v in data.items()}
    except Exception as e:
        log.warning("Could not read neet_answer_key.json: %s", e)
        return None


def parse_neet(soup):
    """
    Parse an NTA NEET response sheet (the `menu-tbl` per-question format).

    Each question is a <table class="menu-tbl"> with label/value rows such as
    'Question ID', 'Option 1 ID' .. 'Option 4 ID', 'Status', 'Chosen Option'.
    The subject is taken from the nearest preceding section heading.
    """
    blocks = soup.find_all("table", class_=re.compile(r"menu-tbl", re.I))
    if not blocks:
        raise ParseError(
            "NEET_NO_QUESTIONS",
            "We could not find any NEET questions in this file.",
            "No <table class='menu-tbl'> blocks found - is this an NTA sheet?")

    answer_key = load_neet_answer_key()
    questions = []
    warnings = []
    UNANSWERED = {"", "-", "--", "not answered", "not attempted",
                  "marked for review", "0"}

    for block in blocks:
        # Build a {label: value} dict from the two-column rows.
        fields = {}
        for tr in block.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) >= 2:
                key = re.sub(r"[^a-z0-9]", "",
                             tds[0].get_text(strip=True).lower())
                fields[key] = tds[1].get_text(strip=True)

        q_id = fields.get("questionid") or fields.get("questionid:")
        if not q_id:
            continue

        option_ids = [fields.get(f"option{n}id", "") for n in range(1, 5)]
        chosen = (fields.get("chosenoption") or "").strip()
        status_txt = (fields.get("status") or "").lower()

        attempted = (chosen.lower() not in UNANSWERED
                     and "not answered" not in status_txt)

        # Subject = nearest preceding heading containing a known subject word.
        subject = "General"
        node = block
        for _ in range(40):
            node = node.find_previous(string=re.compile(
                r"physics|chemistry|botany|zoology|biology", re.I))
            if node:
                subject = normalise_subject(str(node))
                break

        status = "unanswered"
        if attempted:
            if answer_key is not None:
                correct = answer_key.get(str(q_id))
                if correct is None:
                    status = "no_key"          # attempted but key missing
                else:
                    # The key may store an option ID or a 1-4 number.
                    chosen_norm = chosen
                    if chosen in ("1", "2", "3", "4"):
                        idx = int(chosen) - 1
                        chosen_norm = option_ids[idx] if idx < len(option_ids) \
                            else chosen
                    status = "correct" if (chosen == correct or
                                           chosen_norm == correct) \
                        else "incorrect"
            else:
                status = "no_key"

        questions.append({
            "question_id": str(q_id),
            "subject": subject,
            "chosen_index": int(chosen) if chosen in ("1", "2", "3", "4") else None,
            "correct_index": None,
            "status": status,
        })

    if not questions:
        raise ParseError(
            "NEET_EMPTY",
            "No valid NEET questions could be read from this sheet.",
            "menu-tbl blocks were present but none contained a Question ID.")

    if answer_key is None:
        warnings.append(
            "No neet_answer_key.json found - showing attempt counts only. "
            "Add the NTA answer key file to compute an exact NEET score.")
    elif any(q["status"] == "no_key" for q in questions):
        missing = sum(1 for q in questions if q["status"] == "no_key")
        warnings.append(
            f"{missing} attempted question(s) are missing from the answer key "
            "and were excluded from the score.")
    return questions, warnings


# ==============================================================================
#  STEP 4  -  SCORING
# ==============================================================================
def score_questions(questions, exam_type):
    """
    Turn a flat list of parsed questions into a per-subject breakdown plus
    overall totals, applying the exam's marking scheme.
    """
    cfg = EXAM_CONFIG[exam_type]
    neg = cfg["negative"]

    breakdown = {}
    for q in questions:
        subj = q["subject"]
        b = breakdown.setdefault(subj, {
            "questions": 0, "attempted": 0, "correct": 0,
            "incorrect": 0, "unanswered": 0, "score": 0.0,
            "weight": cfg["weights"].get(subj, cfg["default_weight"]),
        })
        b["questions"] += 1
        w = b["weight"]
        st = q["status"]

        if st == "correct" or st == "cancelled":
            b["attempted"] += 1 if st == "correct" else 0
            b["correct"] += 1
            b["score"] += w
        elif st == "incorrect":
            b["attempted"] += 1
            b["incorrect"] += 1
            b["score"] -= neg
        elif st == "unanswered":
            b["unanswered"] += 1
        elif st == "no_key":            # attempted, correctness unknown
            b["attempted"] += 1

    # Tidy numbers: integers stay integers, add per-subject max.
    for b in breakdown.values():
        b["max"] = b["questions"] * b["weight"]
        b["score"] = int(b["score"]) if float(b["score"]).is_integer() \
            else round(b["score"], 2)

    total_q = sum(b["questions"] for b in breakdown.values())
    total_correct = sum(b["correct"] for b in breakdown.values())
    total_incorrect = sum(b["incorrect"] for b in breakdown.values())
    total_unanswered = sum(b["unanswered"] for b in breakdown.values())
    total_attempted = sum(b["attempted"] for b in breakdown.values())
    total_score = sum(b["score"] for b in breakdown.values())
    max_score = sum(b["max"] for b in breakdown.values())

    has_no_key = any(q["status"] == "no_key" for q in questions)
    accuracy = round(100 * total_correct / total_attempted, 1) \
        if total_attempted else 0.0

    return {
        "breakdown": breakdown,
        "total_questions": total_q,
        "attempted": total_attempted,
        "correct": total_correct,
        "incorrect": total_incorrect,
        "unanswered": total_unanswered,
        "total_score": None if has_no_key else (
            int(total_score) if float(total_score).is_integer()
            else round(total_score, 2)),
        "max_score": max_score,
        "accuracy": accuracy,
        "marking_scheme": cfg["scheme"],
    }


# ==============================================================================
#  STEP 5  -  ADVICE  (Gemini when available, deterministic fallback otherwise)
# ==============================================================================
def rule_based_advice(result, exam_type, category):
    """A solid, deterministic analysis used whenever the AI is unavailable."""
    bd = result["breakdown"]
    score = result["total_score"]
    mx = result["max_score"]

    if score is None:
        return ("Your response sheet was parsed successfully, but a NEET score "
                "needs the official NTA answer key. Add neet_answer_key.json to "
                "the backend folder to unlock exact marks and college predictions.")

    ranked = sorted(bd.items(),
                    key=lambda kv: kv[1]["correct"] / max(kv[1]["questions"], 1))
    weakest = ranked[0][0] if ranked else "—"
    strongest = ranked[-1][0] if ranked else "—"
    pct = round(100 * score / mx, 1) if mx else 0

    if exam_type == "MHT-CET":
        if pct >= 90:
            tier = ("an elite range - COEP, VJTI and PICT CSE are realistically "
                    "in reach for the " + category + " category")
        elif pct >= 78:
            tier = ("a strong range - core branches at VJTI/SPIT and most "
                    "branches at PICT are competitive")
        elif pct >= 60:
            tier = ("a mid range - target solid branches at PICT, VIT Pune and "
                    "Cummins rather than the top CSE seats")
        else:
            tier = ("a range where good tier-2 colleges and CAP round strategy "
                    "matter more than chasing the marquee institutes")
    else:
        if pct >= 85:
            tier = "a competitive NEET range for government MBBS counselling"
        elif pct >= 65:
            tier = "a range where state-quota and private MBBS/BDS seats are realistic"
        else:
            tier = "a range where category counselling strategy is decisive"

    return (f"You scored {score}/{mx} ({pct}%), which puts you in {tier}. "
            f"Your strongest subject is {strongest} and {weakest} is dragging "
            f"the total down - that is where focused revision converts fastest "
            f"into rank. Treat this estimate as a planning baseline and confirm "
            f"against this year's official cutoffs before locking preferences.")


def generate_advice(result, exam_type, category):
    if not _ai_client: return rule_based_advice(result, exam_type, category), False
    bd = result.get("breakdown", {})
    subj_lines = "\n".join(f"- {s}: Score {b['score']}/{b['max']}" for s, b in bd.items()) if bd else "Manual Total Score Only"

    prompt = f"""
    Score: {result['total_score']}/{result['max_score']}. Category: {category}. Subjects: {subj_lines}
    You are a brutally honest MHT-CET counselor. Provide detailed analysis. Format in exactly 3 paragraphs separated by blank lines. NO markdown asterisks or lists.
    Paragraph 1: Analyze total score and percentile expectation.
    Paragraph 2: Identify strongest/weakest subjects. List specific realistic Maharashtra colleges (COEP, VJTI, PICT, etc.) for their category.
    Paragraph 3: CAP counseling strategy.
    """
    text = _call_gemini(prompt, max_tokens=600, temperature=0.7)
    return (text, True) if text else (rule_based_advice(result, exam_type, category), False)
# ==============================================================================
#  DATABASE PATHS (Strictly Separated!)
# ==============================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

CET_DB_PATH  = os.path.join(BASE_DIR, "MHTCET_Master.db")
DB_PATH      = CET_DB_PATH
USER_DB_PATH = os.path.join(BASE_DIR, "Users.db") # Cloud-safe path!


def db_status():
    """Quick health check for the SQLite database (used by /health + /predict)."""
    if not os.path.exists(CET_DB_PATH):
        return {"ok": False, "reason": "MHTCET_Master.db not found. "
                "Run py.py then add_colleges.py to build it."}
    try:
        conn = sqlite3.connect(CET_DB_PATH)
        cur = conn.cursor()
        tables = {r[0] for r in cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        out = {"ok": True, "tables": sorted(tables)}
        for t in ("cutoffs", "colleges"):
            if t in tables:
                out[f"{t}_rows"] = cur.execute(
                    f"SELECT COUNT(*) FROM {t}").fetchone()[0]
        out["has_shift_stats"] = "shift_stats" in tables
        conn.close()
        if "cutoffs" not in tables or "colleges" not in tables:
            out["ok"] = False
            out["reason"] = "Database exists but is missing the "  \
                             "'cutoffs' or 'colleges' table."
        return out
    except Exception as e:
        return {"ok": False, "reason": f"DB open failed: {e}"}



# ==============================================================================
#  USER SAAS DATABASE & TOKEN ECONOMY
# ==============================================================================

def init_user_db():
    """Initializes the SaaS User Database and safely migrates new columns."""
    conn = sqlite3.connect(USER_DB_PATH)
    conn.execute('pragma journal_mode=wal')   # concurrent read+write
    cur = conn.cursor()

    # Core table — created fresh on first run
    cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            email           TEXT PRIMARY KEY,
            name            TEXT,
            picture         TEXT,
            is_pro          BOOLEAN  DEFAULT 0,
            tokens_left     INTEGER  DEFAULT 0,
            locked_marks    REAL     DEFAULT NULL,
            my_refer_code   TEXT     UNIQUE,
            referral_count  INTEGER  DEFAULT 0,
            referred_by     TEXT     DEFAULT NULL,
            reward_claimed  INTEGER  DEFAULT 0,
            created_at      TEXT     DEFAULT (datetime('now')),
            last_login      TEXT
        )
    ''')

    # FIX A10: Migration-safe — add new columns to existing DBs without error
    new_cols = [
        ("reward_claimed", "INTEGER DEFAULT 0"),
        ("created_at",     "TEXT DEFAULT (datetime('now'))"),
        ("last_login",     "TEXT"),
    ]
    existing = {row[1] for row in cur.execute("PRAGMA table_info(users)").fetchall()}
    for col, definition in new_cols:
        if col not in existing:
            cur.execute(f"ALTER TABLE users ADD COLUMN {col} {definition}")
            log.info("DB migration: added column '%s'", col)

    # FIX I3: ledger of orders we've already granted, so a payment can never be
    # granted twice (e.g. webhook + /verify-and-grant both firing for one top-up).
    cur.execute('''
        CREATE TABLE IF NOT EXISTS processed_orders (
            order_id   TEXT PRIMARY KEY,
            email      TEXT,
            granted_at TEXT DEFAULT (datetime('now'))
        )
    ''')

    # ── NEW (OTP rebuild): student prediction ledger ──────────────────────────
    # Every prediction run writes one row here. This is the SuperAdmin export
    # source: who searched, with which phone, percentile, and exactly which
    # settings they used. Independent of the `users` economy table above.
    cur.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            name            TEXT,
            phone           TEXT,
            email           TEXT,
            percentile      REAL,
            category        TEXT,
            branch          TEXT,
            city            TEXT,
            shift           TEXT,
            cap_round       TEXT,
            total_matches   INTEGER,
            created_at      TEXT DEFAULT (datetime('now'))
        )
    ''')

    conn.commit()
    conn.close()

# Run once at startup
init_user_db()

def generate_referral_code(name):
    """Generates a unique 6-character code based on the user's name (e.g., ADI9X2)"""
    prefix = re.sub(r'[^A-Z]', '', str(name).upper())[:3]
    if len(prefix) < 3: prefix = (prefix + "GMC")[:3]
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=3))
    return f"{prefix}{suffix}"

# --- Endpoint 1: Google Login Sync ---
@app.route("/sync-user", methods=["POST", "OPTIONS"])
def sync_user():
    if request.method == "OPTIONS": return ("", 204)

    data = request.get_json(silent=True) or {}
    email = data.get("email")
    name = data.get("name", "Student")
    picture = data.get("picture", "")

    if not email:
        return jsonify({"result": "error", "error": "Email required"}), 400

    conn = sqlite3.connect(USER_DB_PATH)  # Use USER DB!
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    user = cur.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()

    if not user:
        refer_code = generate_referral_code(name)
        while cur.execute("SELECT 1 FROM users WHERE my_refer_code = ?", (refer_code,)).fetchone():
            refer_code = generate_referral_code(name)

        cur.execute('''
            INSERT INTO users (email, name, picture, my_refer_code)
            VALUES (?, ?, ?, ?)
        ''', (email, name, picture, refer_code))
        conn.commit()
        user = cur.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
    else:
        # Update picture + last_login on every sign-in
        cur.execute(
            "UPDATE users SET picture = ?, last_login = datetime('now') WHERE email = ?",
            (picture, email)
        )
        conn.commit()
        user = cur.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()

    conn.close()

    return jsonify({
        "result": "success",
        "user": {
            "email": user["email"],
            "name": user["name"],
            "is_pro": bool(user["is_pro"]),
            "tokens_left": user["tokens_left"],
            "locked_marks": user["locked_marks"],
            "my_refer_code": user["my_refer_code"],
            "referral_count": user["referral_count"],
            "reward_claimed": user["reward_claimed"] if "reward_claimed" in user.keys() else 0,
            "created_at": user["created_at"] if "created_at" in user.keys() else None,
        }
    })

# --- NEW (OTP rebuild): Phone Sign-in Sync -----------------------------------
# Frontend verifies the OTP via Firebase, then calls this with name + phone.
# We synthesize a stable email key (phone@phone.gmc) so the ENTIRE existing
# backend (tokens, is_pro, redeem-key, admin/stats) keeps working untouched.
# New users start with tokens_left = FREE_PREDICTIONS (4 free runs).
FREE_PREDICTIONS = 4

def phone_to_email(phone: str) -> str:
    """Stable identity key from a phone number. e.g. '9876543210' -> '9876543210@phone.gmc'"""
    digits = re.sub(r"\D", "", str(phone or ""))
    return f"{digits}@phone.gmc" if digits else ""

@app.route("/sync-phone", methods=["POST", "OPTIONS"])
def sync_phone():
    if request.method == "OPTIONS": return ("", 204)

    data  = request.get_json(silent=True) or {}
    name  = (data.get("name")  or "Student").strip()
    phone = re.sub(r"\D", "", str(data.get("phone") or ""))
    picture = data.get("picture", "")   # auto-generated avatar URL from frontend

    if len(phone) < 10:
        return jsonify({"result": "error", "error": "A valid phone number is required."}), 400

    email = phone_to_email(phone)

    conn = sqlite3.connect(USER_DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    user = cur.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()

    if not user:
        refer_code = generate_referral_code(name)
        while cur.execute("SELECT 1 FROM users WHERE my_refer_code = ?", (refer_code,)).fetchone():
            refer_code = generate_referral_code(name)
        # NOTE: tokens_left seeded to FREE_PREDICTIONS so phone users get 4 free runs.
        cur.execute('''
            INSERT INTO users (email, name, picture, my_refer_code, tokens_left, last_login)
            VALUES (?, ?, ?, ?, ?, datetime('now'))
        ''', (email, name, picture, refer_code, FREE_PREDICTIONS))
        conn.commit()
        user = cur.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
    else:
        # Returning user: refresh name/picture/last_login, keep economy intact.
        cur.execute(
            "UPDATE users SET name = ?, picture = ?, last_login = datetime('now') WHERE email = ?",
            (name, picture, email)
        )
        conn.commit()
        user = cur.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()

    conn.close()

    return jsonify({
        "result": "success",
        "user": {
            "email": user["email"],
            "name": user["name"],
            "picture": user["picture"],
            "phone": phone,
            "is_pro": bool(user["is_pro"]),
            "tokens_left": user["tokens_left"],
            "my_refer_code": user["my_refer_code"],
            "referral_count": user["referral_count"],
            "created_at": user["created_at"] if "created_at" in user.keys() else None,
        }
    })
@app.route("/get-user", methods=["GET", "OPTIONS"])
def get_user():
    """Lightweight state refresh — frontend calls this on every page load."""
    if request.method == "OPTIONS": return ("", 204)
    email = request.args.get("email", "").strip()
    if not email:
        return jsonify({"result": "error", "error": "Email required"}), 400

    conn = sqlite3.connect(USER_DB_PATH)
    conn.row_factory = sqlite3.Row
    user = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
    conn.close()

    if not user:
        return jsonify({"result": "error", "error": "User not found"}), 404

    return jsonify({
        "result": "success",
        "user": {
            "email": user["email"],
            "name": user["name"],
            "picture": user["picture"],
            "is_pro": bool(user["is_pro"]),
            "tokens_left": user["tokens_left"],
            "locked_marks": user["locked_marks"],
            "my_refer_code": user["my_refer_code"],
            "referral_count": user["referral_count"],
            "reward_claimed": user["reward_claimed"] if "reward_claimed" in user.keys() else 0,
            "created_at": user["created_at"] if "created_at" in user.keys() else None,
        }
    })

# --- Endpoint 2: Use a Token & Lock Marks ---
@app.route("/use-token", methods=["POST", "OPTIONS"])
def use_token():
    if request.method == "OPTIONS": return ("", 204)

    data = request.get_json(silent=True) or {}
    email = data.get("email")
    current_marks = float(data.get("marks", 0))

    if not email: return jsonify({"allowed": False, "reason": "NOT_LOGGED_IN", "tokens_left": 0})

    conn = sqlite3.connect(USER_DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    user = cur.execute("SELECT is_pro, tokens_left, locked_marks FROM users WHERE email = ?", (email,)).fetchone()

    if not user or not user["is_pro"]:
        conn.close()
        return jsonify({"allowed": False, "reason": "NOT_PRO", "tokens_left": 0})

    tokens_left  = user["tokens_left"]
    locked_marks = user["locked_marks"]

    if tokens_left <= 0:
        if locked_marks is not None and round(current_marks, 2) != round(float(locked_marks), 2):
            conn.close()
            return jsonify({
                "allowed": False,
                "reason": "TOKENS_EMPTY",
                "tokens_left": 0,
                "locked_marks": locked_marks,
                "message": "Your marks are locked. Top up tokens to change them."
            })
        # Same marks — allow re-filtering without burning a token
        conn.close()
        return jsonify({
            "allowed": True,
            "tokens_left": 0,
            "warning": "MARKS_LOCKED",
            "message": "Showing results for your locked marks. Top up to change them."
        })

    new_tokens = tokens_left - 1
    new_locked  = round(current_marks, 2) if new_tokens == 0 else locked_marks

    cur.execute("UPDATE users SET tokens_left = ?, locked_marks = ? WHERE email = ?",
                (new_tokens, new_locked, email))
    conn.commit()
    conn.close()

    return jsonify({"allowed": True, "tokens_left": new_tokens})


def _do_grant_pro(email: str, promo_code: str, is_topup: bool = False, order_id: str = None):
    """
    Internal helper — upgrades the user to PRO and rewards the referrer.
    Called only AFTER payment signature is verified.

    FIX I3: if order_id is supplied and was already processed, this is a no-op,
    so webhook + verify-and-grant can't double-grant (especially top-up tokens).
    Returns True if a grant happened, False if it was a duplicate.
    """
    conn = sqlite3.connect(USER_DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if order_id:
        already = cur.execute(
            "SELECT 1 FROM processed_orders WHERE order_id = ?", (order_id,)
        ).fetchone()
        if already:
            conn.close()
            log.info("Skipping duplicate grant for order=%s email=%s", order_id, email)
            return False
        cur.execute(
            "INSERT OR IGNORE INTO processed_orders (order_id, email) VALUES (?, ?)",
            (order_id, email)
        )

    if is_topup:
        # Top-up: add 3 more tokens, keep existing PRO, clear locked marks
        cur.execute(
            "UPDATE users SET tokens_left = tokens_left + 3, locked_marks = NULL WHERE email = ?",
            (email,)
        )
    else:
        # Fresh purchase: grant PRO, reset tokens to 3, clear locked marks
        cur.execute(
            "UPDATE users SET is_pro = 1, tokens_left = 3, locked_marks = NULL WHERE email = ?",
            (email,)
        )

    # Reward the referrer (only on fresh purchase, not top-up)
    if promo_code and not is_topup:
        user = cur.execute("SELECT referred_by FROM users WHERE email = ?", (email,)).fetchone()
        if user and not user["referred_by"]:
            cur.execute("UPDATE users SET referred_by = ? WHERE email = ?", (promo_code, email))
            cur.execute(
                "UPDATE users SET referral_count = referral_count + 1 WHERE my_refer_code = ?",
                (promo_code,)
            )

    conn.commit()
    conn.close()
    return True


# --- SECURE: Verify with Cashfree → THEN Grant PRO ---
@app.route("/verify-and-grant", methods=["POST", "OPTIONS"])
def verify_and_grant():
    """
    Called by premium.html after Cashfree checkout returns.
    Confirms the order is actually PAID with Cashfree before granting PRO.
    Payload: { order_id, email, promo_code }
    """
    if request.method == "OPTIONS": return ("", 204)
    data = request.get_json(silent=True) or {}

    order_id    = data.get("order_id", "").strip()
    email       = data.get("email", "").strip()
    promo_code  = data.get("promo_code", "").strip().upper()

    if not all([order_id, email]):
        return jsonify({"result": "error", "error": "Missing required fields."}), 400

    # --- Ask Cashfree for the real, server-side order status ---
    try:
        paid = _cf_order_is_paid(order_id)
    except Exception as e:
        log.error("Cashfree verify error for %s: %s", email, e)
        return jsonify({"result": "error", "error": "Verification failed."}), 502

    if not paid:
        log.warning("⚠️  PAYMENT NOT PAID for email=%s  order=%s", email, order_id)
        return jsonify({"result": "error", "error": "Payment verification failed."}), 403

    # --- Payment confirmed → determine if this is a top-up ---
    conn = sqlite3.connect(USER_DB_PATH)
    conn.row_factory = sqlite3.Row
    user = conn.execute("SELECT is_pro FROM users WHERE email = ?", (email,)).fetchone()
    conn.close()

    is_topup = bool(user and user["is_pro"])   # already PRO = this is a top-up

    _do_grant_pro(email, promo_code, is_topup, order_id=order_id)
    log.info("✅  PRO granted  email=%s  topup=%s  promo=%s", email, is_topup, promo_code or "—")
    return jsonify({"result": "success", "is_topup": is_topup})


# --- Verify a one-off COUNSELLING payment (called by predict.html) ---
@app.route("/verify-payment", methods=["POST", "OPTIONS"])
def verify_payment():
    """
    Called by predict.html after a counselling-fee Cashfree checkout.
    This is a one-time payment (not a PRO subscription), so we only confirm
    the order is actually PAID — we do NOT grant PRO tokens here.
    Payload: { order_id, email }
    """
    if request.method == "OPTIONS": return ("", 204)
    data = request.get_json(silent=True) or {}

    order_id = data.get("order_id", "").strip()
    email    = data.get("email", "").strip()

    if not order_id:
        return jsonify({"result": "error", "error": "Missing order_id."}), 400

    try:
        paid = _cf_order_is_paid(order_id)
    except Exception as e:
        log.error("Cashfree verify-payment error for %s: %s", email, e)
        return jsonify({"result": "error", "error": "Verification failed."}), 502

    if not paid:
        log.warning("⚠️  COUNSELLING PAYMENT NOT PAID  email=%s  order=%s", email, order_id)
        return jsonify({"result": "error", "error": "Payment not completed or failed."}), 403

    log.info("✅  Counselling payment confirmed  email=%s  order=%s", email, order_id)
    return jsonify({"result": "success", "order_id": order_id})


# --- Verify a one-off BOOKING payment (called by booking.html) ---
@app.route("/verify-booking-payment", methods=["POST", "OPTIONS"])
def verify_booking_payment():
    """
    Called by booking.html after a 1:1 session Cashfree checkout.
    Confirms the order is actually PAID before the frontend syncs the booking
    to the Google Sheet. One-time payment, so no PRO grant here.
    Payload: { order_id, email }
    """
    if request.method == "OPTIONS": return ("", 204)
    data = request.get_json(silent=True) or {}

    order_id = data.get("order_id", "").strip()
    email    = data.get("email", "").strip()

    if not order_id:
        return jsonify({"result": "error", "error": "Missing order_id."}), 400

    try:
        paid = _cf_order_is_paid(order_id)
    except Exception as e:
        log.error("Cashfree verify-booking error for %s: %s", email, e)
        return jsonify({"result": "error", "error": "Verification failed."}), 502

    if not paid:
        log.warning("⚠️  BOOKING PAYMENT NOT PAID  email=%s  order=%s", email, order_id)
        return jsonify({"result": "error", "error": "Payment not completed or failed."}), 403

    log.info("✅  Booking payment confirmed  email=%s  order=%s", email, order_id)
    return jsonify({"result": "success", "order_id": order_id})


# --- DEPRECATED: /grant-pro (kept so old JS doesn't 404, but refuses all grants) ---
# --- Endpoint: Grant Pro & Reward Referrer (SECURED) ---
@app.route("/grant-pro", methods=["POST", "OPTIONS"])
def grant_pro():
    if request.method == "OPTIONS": return ("", 204)
    data = request.get_json(silent=True) or {}
    email = data.get("email")
    promo_code = data.get("promo_code", "").strip().upper()

    # 🚨 Verify the order is truly PAID with Cashfree so payments can't be spoofed
    order_id = data.get("order_id")

    if not order_id:
        return jsonify({"result": "error", "error": "Missing order_id."}), 400

    try:
        if not _cf_order_is_paid(order_id):
            log.error(f"HACK ATTEMPT BLOCKED for {email}: order not PAID.")
            return jsonify({"result": "error", "error": "Payment not verified. Hack attempt blocked."}), 400
    except Exception as e:
        log.error(f"Cashfree verify error for {email}: {e}")
        return jsonify({"result": "error", "error": "Verification failed."}), 502

    # --- Proceed to grant PRO ---
    conn = sqlite3.connect(USER_DB_PATH)
    cur = conn.cursor()

    # Give the buyer PRO, reset tokens to 3, and wipe locked marks
    cur.execute("UPDATE users SET is_pro = 1, tokens_left = 3, locked_marks = NULL WHERE email = ?", (email,))

    # If a promo code was used, reward the referrer!
    if promo_code:
        user = cur.execute("SELECT referred_by FROM users WHERE email = ?", (email,)).fetchone()
        if user and not user[0]: # If they haven't used a code before
            cur.execute("UPDATE users SET referred_by = ? WHERE email = ?", (promo_code, email))
            cur.execute("UPDATE users SET referral_count = referral_count + 1 WHERE my_refer_code = ?", (promo_code,))

    conn.commit()
    conn.close()
    return jsonify({"result": "success"})
# --- Category mapping -------------------------------------------------------
# Maps the user's (caste) selection to a reservation "relaxation" in percentile
# points relative to the General-Open (GOPEN) cut-off. These are realistic
# MHT-CET averages; when category-specific columns exist in the DB they are
# preferred, otherwise this offset approximates the reserved cut-off.
CATEGORY_RELAXATION = {
    "OPEN":            0.0,
    "EWS":             1.5,
    "TFWS":           -1.0,
    "OBC":             3.0,
    "SEBC":            3.5,
    "VJ":              6.0,
    "NT1":             6.5,
    "NT2":             6.0,
    "NT3":             5.5,
    "SC":              9.0,
    "ST":             14.0,
    "JAIN_MINORITY":   0.0,
    "RELIGIOUS_MINORITY":  0.0,
    "LINGUISTIC_MINORITY":  0.0,
    "OTHER_MINORITY":  0.0,
    "PWD":             8.0,
    "DEFENCE":         5.0,
    # --- RED TEAM FIX: Prevent Sub-Categories from defaulting to OPEN ---
    "GUJARATI_LINGUISTIC": 0.0,
    "HINDI_LINGUISTIC": 0.0,
    "SINDHI_LINGUISTIC": 0.0,
    "URDU_LINGUISTIC": 0.0,
    "TELUGU_LINGUISTIC": 0.0,
    "KANNADA_LINGUISTIC": 0.0,
    "CHRISTIAN_MINORITY": 0.0,
    "SIKH_MINORITY": 0.0,
    "PARSI_MINORITY": 0.0,
    "BUDDHIST_MINORITY": 0.0,
    "MUSLIM_MINORITY": 0.0,
}

# Friendly label -> internal key (frontend may send either).
CATEGORY_ALIASES = {
    "general": "OPEN", "open": "OPEN",
    "ews": "EWS", "tfws": "TFWS",
    "obc": "OBC", "sebc": "SEBC",
    "vj": "VJ", "vjnt": "VJ", "dtvj": "VJ",
    "nt1": "NT1", "nt-b": "NT1", "ntb": "NT1",
    "nt2": "NT2", "nt-c": "NT2", "ntc": "NT2",
    "nt3": "NT3", "nt-d": "NT3", "ntd": "NT3",
    "sc": "SC", "st": "ST",
    "jain": "JAIN_MINORITY", "jain minority": "JAIN_MINORITY",
    "religious": "RELIGIOUS_MINORITY", "religious minority": "RELIGIOUS_MINORITY",     # <-- ADDED
    "linguistic": "LINGUISTIC_MINORITY", "linguistic minority": "LINGUISTIC_MINORITY", # <-- ADDED
    "minority": "OTHER_MINORITY", "other minority": "OTHER_MINORITY",
    "pwd": "PWD", "defence": "DEFENCE", "defense": "DEFENCE",
}


def normalise_category(raw):
    if not raw:
        return "OPEN"
    key = str(raw).strip().lower()
    return CATEGORY_ALIASES.get(key, raw.strip().upper()
                                if raw.strip().upper() in CATEGORY_RELAXATION
                                else "OPEN")


# --- Branch mapping ---------------------------------------------------------
BRANCH_WILDCARDS = {
    "CSE":   "%Computer%",
    "CE":    "%Computer%",
    "IT":    "%Information%",
    "AIDS":  "%Artificial%",
    "AIML":  "%Machine Learning%",
    "DS":    "%Data Science%",
    "ENTC":  "%Electronics and Telecommunication%",
    "ECE":   "%Electronics%",
    "EE":    "%Electrical%",
    "ME":    "%Mechanical%",
    "CIVIL": "%Civil%",
    "CHEM":  "%Chemical%",
    "ROBO":  "%Robotics%",
}

# --- City aliases -----------------------------------------------------------
# Institute names in the MHT-CET data spell some cities inconsistently
# (Nashik/Nasik, Aurangabad/Chh. Sambhajinagar, Mumbai/Bombay, etc.).
# When the user picks a city we match against ALL of its known spellings so
# the filter doesn't silently drop valid colleges or leak the wrong ones.
CITY_ALIASES = {
    "pune":        ["pune", "pimpri", "chinchwad", "pcmc", "lonavala", "talegaon", "lonavla"],
    "mumbai":      ["mumbai", "bombay", "andheri", "bandra", "borivali", "powai", "matunga", "vidyavihar", "wadala"],
    "navi mumbai": ["navi mumbai", "vashi", "panvel", "kharghar", "nerul", "belapur", "airoli"],
    "thane":       ["thane", "kalyan", "dombivli", "bhiwandi", "ulhasnagar", "ambernath"],
    "nashik":      ["nashik", "nasik"],
    "nagpur":      ["nagpur"],
    "aurangabad":  ["aurangabad", "sambhajinagar", "chh. sambhajinagar", "chhatrapati sambhajinagar"],
    "amravati":    ["amravati"],
    "kolhapur":    ["kolhapur"],
    "solapur":     ["solapur", "sholapur"],
    "jalgaon":     ["jalgaon"],
    "nanded":      ["nanded"],
    "latur":       ["latur"],
    "sangli":      ["sangli", "miraj"],
    "satara":      ["satara", "karad"],
    "ahmednagar":  ["ahmednagar", "ahilyanagar", "a'nagar"],
    "akola":       ["akola"],
    "chandrapur":  ["chandrapur"],
    "dhule":       ["dhule"],
    "ratnagiri":   ["ratnagiri"],
    "raigad":      ["raigad", "alibag", "mahad"],
    "wardha":      ["wardha"],
    "yavatmal":    ["yavatmal"],
    "sindhudurg":  ["sindhudurg", "kankavli"],
    "palghar":     ["palghar", "vasai", "virar", "boisar"],
}


def city_keywords(region_str):
    """
    Turns the user's comma-separated city string into a flat list of
    lowercase search keywords (expanded via CITY_ALIASES). Unknown cities
    fall back to the raw city name so nothing is lost.
    """
    out = []
    for raw in (region_str or "").split(","):
        c = raw.strip().lower()
        if not c:
            continue
        out.extend(CITY_ALIASES.get(c, [c]))
    return out


# --- Shift normalisation ----------------------------------------------------
def load_shift_stats(conn):
    """
    Returns {shift_id: mean_percentile} from the optional 'shift_stats' table.
    If the table is absent we return {} and normalisation becomes a no-op.
    """
    try:
        cur = conn.cursor()
        rows = cur.execute(
            "SELECT Shift, Mean_Percentile FROM shift_stats").fetchall()
        return {str(r[0]).strip().upper(): float(r[1]) for r in rows if r[1]}
    except Exception:
        return {}


def normalise_percentile(raw_pct, shift, shift_stats):
    """
    Shift-difficulty correction. A student in a HARD shift (low mean percentile)
    gets a small upward correction so they compare fairly against cut-offs that
    were themselves set in a mix of shifts.

        difficulty_index = global_mean - shift_mean      (hard shift -> +ve)
        normalised       = raw + difficulty_index * 0.5  (capped at +/-1.5)
    """
    if not shift_stats or not shift:
        return raw_pct, 0.0
    shift_key = str(shift).strip().upper()
    if shift_key not in shift_stats:
        return raw_pct, 0.0
    global_mean = sum(shift_stats.values()) / len(shift_stats)
    diff = global_mean - shift_stats[shift_key]
    correction = max(-1.5, min(1.5, diff * 0.5))
    return round(min(100.0, raw_pct + correction), 4), round(correction, 3)


# --- Trend projection -------------------------------------------------------
def project_cutoff(year_cutoffs):
    """
    year_cutoffs: {year(int): percentile(float)} for a single choice code.
    Applies a weighted moving average that favours recent years, then projects
    one year forward using the recent delta (velocity).
    Returns (projected_cutoff, trend_label, latest_known).
    """
    if not year_cutoffs:
        return None, "unknown", None

    years = sorted(year_cutoffs.keys())
    latest = year_cutoffs[years[-1]]

    if len(years) == 1:
        return latest, "flat", latest

    # Velocity = average year-on-year change, recent change weighted higher.
    deltas = []
    for i in range(1, len(years)):
        gap = years[i] - years[i - 1]
        if gap > 0:
            deltas.append((year_cutoffs[years[i]] - year_cutoffs[years[i - 1]])
                           / gap)
    if not deltas:
        return latest, "flat", latest

    # weight recent deltas more (linear weights 1,2,3,...)
    w = list(range(1, len(deltas) + 1))
    velocity = sum(d * wi for d, wi in zip(deltas, w)) / sum(w)
    velocity = max(-2.0, min(2.0, velocity))     # clamp wild swings

    projected = round(min(100.0, latest + velocity), 3)
    if velocity > 0.15:
        trend = "rising"
    elif velocity < -0.15:
        trend = "falling"
    else:
        trend = "stable"
    return projected, trend, latest


# --- Risk + confidence ------------------------------------------------------
def risk_and_confidence(student_pct, projected_cutoff):
    """
    margin = student_pct - projected_cutoff
    Returns (risk_label, chance_percent 0-100).
    chance_percent uses a logistic curve centred on the cut-off.
    """
    margin = student_pct - projected_cutoff
    # logistic: steepness k tuned so +1.5 margin ~= 92%, -1.5 ~= 8%
    chance = 100.0 / (1.0 + math.exp(-1.6 * margin))
    chance = round(max(1.0, min(99.0, chance)), 1)

    if margin >= 1.0:
        risk = "Safe"
    elif margin >= -0.5:
        risk = "Moderate"
    elif margin >= -2.5:
        risk = "Reach"
    else:
        risk = "Unlikely"
    return risk, chance


def percentile_from_rank(rank):
    """Rough rank->percentile fallback (MHT-CET ~4.5 lakh candidates)."""
    try:
        rank = float(rank)
    except Exception:
        return None
    total = 450000.0
    if rank <= 0:
        return None
    pct = 100.0 * (1.0 - (rank / total))
    return round(max(1.0, min(99.999, pct)), 4)


# ── NEW (OTP rebuild): persist one student prediction row ─────────────────────
def _save_student_row(data, meta, matches):
    """Write a single students-table row from a completed prediction."""
    email = (data.get("email") or "").strip()
    phone = re.sub(r"\D", "", str(data.get("phone") or ""))
    if not phone and email.endswith("@phone.gmc"):
        phone = email.split("@")[0]
    name = (data.get("name") or "").strip()

    # If name missing in the predict payload, recover it from the users table.
    if not name and email:
        try:
            c = sqlite3.connect(USER_DB_PATH); c.row_factory = sqlite3.Row
            row = c.execute("SELECT name FROM users WHERE email = ?", (email,)).fetchone()
            c.close()
            if row: name = row["name"]
        except Exception:
            pass

    category = (meta or {}).get("category") or (data.get("category") or "")
    branch   = (meta or {}).get("branch_pref") or (data.get("branch") or "ALL")
    city     = (data.get("location") or data.get("city") or data.get("region") or "")
    shift    = (data.get("shift") or "")
    cap_rd   = str((meta or {}).get("cap_round") or data.get("cap_round") or "ALL")
    pct      = (meta or {}).get("raw_percentile")
    if pct is None:
        try: pct = float(data.get("percentile"))
        except Exception: pct = None

    conn = sqlite3.connect(USER_DB_PATH)
    conn.execute('''
        INSERT INTO students
          (name, phone, email, percentile, category, branch, city, shift, cap_round, total_matches)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (name, phone, email, pct, category, branch, city, shift, cap_rd, len(matches or [])))
    conn.commit()
    conn.close()


def run_prediction(user):
    """
    Core predictor. `user` is a dict from the frontend form.
    Returns (matches list, meta dict).  Raises ParseError on bad input / no DB.
    """
    status = db_status()
    if not status["ok"]:
        raise ParseError("DB_UNAVAILABLE",
                          "The college database is not ready yet.",
                          status.get("reason", "unknown"), status=503)

    # ---- resolve the student's effective percentile ------------------------
    percentile = user.get("percentile")
    rank = user.get("rank")
    try:
        percentile = float(percentile) if percentile not in (None, "") else None
    except Exception:
        percentile = None
    if percentile is None and rank not in (None, ""):
        percentile = percentile_from_rank(rank)
    if percentile is None:
        raise ParseError("NO_SCORE",
                          "Please enter your percentile or your rank.",
                          "Neither percentile nor rank supplied.")
    percentile = max(0.0, min(100.0, percentile))

    category = normalise_category(user.get("category") or user.get("caste"))
    relaxation = CATEGORY_RELAXATION.get(category, 0.0)
    branch_pref = (user.get("branch") or "").strip().upper()
    shift = (user.get("shift") or "").strip()
    region = (user.get("location") or user.get("region") or "").strip()
    # Expand the chosen cities into alias keywords once (used as a hard filter).
    city_kws = city_keywords(region)

    # CAP round filter: "1", "2", "3" to restrict to a single round, or
    # "ALL"/"" to use every round (default).
    cap_round = str(user.get("cap_round") or user.get("round") or "ALL").strip().upper()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # ---- shift normalisation ----------------------------------------------
    shift_stats = load_shift_stats(conn)
    norm_pct, shift_corr = normalise_percentile(percentile, shift, shift_stats)

    # "effective" percentile the reserved student competes with on OPEN seats
    effective_pct = min(99.99, norm_pct + relaxation)

    # ---- pull every choice code with its multi-year cut-off history --------
    # We look for branches whose latest cut-off is within reach (effective+2.5).
    sql = """
        SELECT c.Choice_Code, c.Institute_Name, c.Course_Name, c.Intake,
               c.Institute_Code, c.Minority_Status,
               t.Year, t.Round, t.Percentile, t.Seat_Type
        FROM colleges c
        JOIN cutoffs  t ON c.Choice_Code = t.Choice_Code
    """
    params = []
    where_clauses = []
    if branch_pref and branch_pref not in ("ALL", "OTHER", ""):
        where_clauses.append("c.Course_Name LIKE ?")
        params.append(BRANCH_WILDCARDS.get(branch_pref, f"%{branch_pref}%"))
    if cap_round in ("1", "2", "3"):
        where_clauses.append("t.Round = ?")
        params.append(int(cap_round))
    if where_clauses:
        sql += " WHERE " + " AND ".join(where_clauses)
    rows = cur.execute(sql, params).fetchall()
    conn.close()

    # ---- group by choice code, keep best (lowest) cut-off per year ---------
    grouped = {}
    for r in rows:
        cc = r["Choice_Code"]
        g = grouped.setdefault(cc, {
            "info": {
                "college": r["Institute_Name"],
                "branch": r["Course_Name"],
                "intake": r["Intake"],
                "inst_code": r["Institute_Code"],
                "minority": r["Minority_Status"],
            },
            "years": {},
            "rounds": set(),
        })
        try:
            yr = int(re.search(r"20\d\d", str(r["Year"])).group())
        except Exception:
            continue
        # remember which CAP rounds this branch appeared in
        try:
            g["rounds"].add(int(r["Round"]))
        except Exception:
            pass
        pct = float(r["Percentile"])
        # keep the representative (highest open) cut-off per year
        if yr not in g["years"] or pct > g["years"][yr]:
            g["years"][yr] = pct

    # ---- score every branch ------------------------------------------------
    # ---- score every branch ------------------------------------------------
    matches = []
    for cc, g in grouped.items():
        projected, trend, latest = project_cutoff(g["years"])
        if projected is None:
            continue

        minority_status_db = str(g["info"]["minority"] or "NA").upper()
        minority_branch = minority_status_db not in ("NA", "", "NONE", "NULL")

        # ---- THE DYNAMIC MINORITY ENGINE ----
        adjusted_projected = projected
        user_cat = category.upper()

        if minority_branch:
            # 1. Jain Minority (Massive cutoff drop for Jain students in Jain colleges like SNJB)
            if "JAIN" in user_cat and "JAIN" in minority_status_db:
                adjusted_projected = max(1.0, projected - 75.0)

            # 2. Linguistic Minority (Gujarati / Hindi / Sindhi)
            elif "LINGUISTIC" in user_cat and any(k in minority_status_db for k in ["LINGUISTIC", "HINDI", "GUJARATI", "SINDHI"]):
                adjusted_projected = max(1.0, projected - 50.0)

            # 3. Religious Minority
            elif "RELIGIOUS" in user_cat and "RELIGIOUS" in minority_status_db:
                adjusted_projected = max(1.0, projected - 40.0)

            # 4. Generic Minority Fallback
            elif "MINORITY" in user_cat and "MINORITY" in minority_status_db:
                adjusted_projected = max(1.0, projected - 40.0)

        # Risk profiling uses the strictly adjusted minority cutoff!
        risk, chance = risk_and_confidence(effective_pct, adjusted_projected)

        if risk == "Unlikely":
            continue            # >2.5 percentile short - not worth listing

        # ---- CITY FILTER (hard) -------------------------------------------
        # If the user picked one or more cities, ONLY keep colleges whose
        # institute name matches one of those cities (alias-expanded). This is
        # a real filter now, not just a soft nudge, so "Pune, Nashik" never
        # leaks colleges from other cities.
        college_name_lower = g["info"]["college"].lower()
        loc_match = False
        if city_kws:
            loc_match = any(k in college_name_lower for k in city_kws)
            if not loc_match:
                continue            # not in a selected city -> drop it

        if loc_match:
            chance = min(99.0, chance + 3.0)

        matches.append({
            "choice_code": cc,
            "college": g["info"]["college"],
            "branch": g["info"]["branch"],
            "intake": g["info"]["intake"],
            "inst_code": g["info"]["inst_code"],
            "minority": g["info"]["minority"],
            "cutoff": round(adjusted_projected, 2), # Display the real minority cutoff!
            "cutoff_latest": round(latest, 2) if latest else None,
            "history": {str(y): round(v, 2)
                        for y, v in sorted(g["years"].items())},
            "trend": trend,
            "risk": risk,
            "chance": chance,
            "location_match": loc_match,
            "minority_branch": minority_branch,
            "rounds": sorted(g["rounds"]),
        })
    # default sort: best chance first, then higher cut-off (better college)
    matches.sort(key=lambda m: (-m["chance"], -m["cutoff"]))

    meta = {
        "raw_percentile": round(percentile, 3),
        "normalised_percentile": round(norm_pct, 3),
        "shift_correction": shift_corr,
        "category": category,
        "category_relaxation": relaxation,
        "effective_percentile": round(effective_pct, 3),
        "shift": shift or None,
        "branch_pref": branch_pref or "ALL",
        "cap_round": cap_round if cap_round in ("1", "2", "3") else "ALL",
        "total_matches": len(matches),
        "buckets": {
            "Safe": sum(1 for m in matches if m["risk"] == "Safe"),
            "Moderate": sum(1 for m in matches if m["risk"] == "Moderate"),
            "Reach": sum(1 for m in matches if m["risk"] == "Reach"),
        },
        "shift_stats_used": bool(shift_stats),
    }
    return matches, meta


# --- Direct Gemini HTTP caller (gemini-2.5-pro, JSON-capable) ---------------
def _call_gemini(prompt, model="gemini-2.5-pro", max_tokens=8192,
                 temperature=0.7, json_mode=False, timeout=60):
    """
    Calls the Gemini REST API directly via urllib. Returns text or None.

    IMPORTANT: gemini-2.5-pro is a THINKING model -- it spends output tokens on
    internal reasoning. max_tokens must be generous (>=4096) or the model
    returns EMPTY text with finishReason=MAX_TOKENS.
    """
    if not GEMINI_KEY:
        log.error("=" * 64)
        log.error(" GEMINI KEY IS EMPTY -- paste your key into GEMINI_KEY (top of file).")
        log.error("=" * 64)
        return None

    url = ("https://generativelanguage.googleapis.com/v1beta/models/"
           f"{model}:generateContent?key={GEMINI_KEY}")

    gen_cfg = {"temperature": temperature, "maxOutputTokens": max_tokens}
    if json_mode:
        gen_cfg["responseMimeType"] = "application/json"

    body = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": gen_cfg,
    }).encode("utf-8")

    try:
        req = urllib.request.Request(
            url, data=body,
            headers={"Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(req, timeout=timeout) as r:
            data = json.loads(r.read())
    except urllib.error.HTTPError as e:
        err = e.read().decode("utf-8", errors="ignore")
        log.error("=" * 64)
        log.error(" GEMINI HTTP %s  (model=%s)", e.code, model)
        log.error(" %s", err[:600])
        if "API key not valid" in err:
            log.error(" >> Your API KEY IS INVALID. Get a new one at "
                      "https://aistudio.google.com/app/apikey")
        elif e.code == 403:
            log.error(" >> 403: enable the 'Generative Language API' for this key.")
        elif e.code == 404:
            log.error(" >> 404: model not found. Try model='gemini-2.5-flash'.")
        elif e.code == 429:
            log.error(" >> 429: quota / rate limit hit. Wait, or upgrade the plan.")
        log.error("=" * 64)
        return None
    except Exception as e:
        log.error(" GEMINI network error: %s: %s", type(e).__name__, e)
        return None

    try:
        cand = (data.get("candidates") or [{}])[0]
        finish = cand.get("finishReason", "?")
        parts = (cand.get("content") or {}).get("parts") or []
        text = "".join(p.get("text", "") for p in parts).strip()
        if not text:
            log.error(" GEMINI returned EMPTY text. finishReason=%s", finish)
            if finish == "MAX_TOKENS":
                log.error(" >> Model spent all tokens thinking. Raise max_tokens.")
            log.error(" Raw response: %s", json.dumps(data)[:500])
            return None
        log.info(" GEMINI (%s) replied OK -- %d chars.", model, len(text))
        return text
    except Exception as e:
        log.error(" GEMINI parse error: %s | raw: %s", e, json.dumps(data)[:500])
        return None

# --- Gemini placement summaries (JSON, gemini-2.5-pro) ----------------------
def gemini_placement_summaries(top_colleges):
    """Returns [{college,branch,avg_package,top_recruiters,summary}, ...]."""
    fallback = [{
        "college": c["college"], "branch": c["branch"],
        "avg_package": "N/A", "top_recruiters": [],
        "summary": ("Gemini call failed -- check the Python terminal for the "
                    "exact error line."),
    } for c in top_colleges]

    if not GEMINI_KEY or not top_colleges:
        return fallback

    listing = "\n".join(f"{i+1}. {c['college']} - {c['branch']}"
                        for i, c in enumerate(top_colleges))
    prompt = (
        "You are a strict Maharashtra engineering placement analyst. For each "
        "college+branch below, give a realistic placement snapshot.\n\n"
        f"{listing}\n\n"
        "Return ONLY a JSON array. Each item must have keys: "
        '"college", "branch", "avg_package" (e.g. "6-9 LPA"), '
        '"top_recruiters" (array of 3 company strings), '
        '"summary" (one honest sentence). '
        "CRITICAL: If you do not have exact data for a specific college, give a realistic estimate based on its tier and add '(Est.)' to the avg_package."
    )

    text = _call_gemini(prompt, model="gemini-2.5-flash",
                        max_tokens=2048, temperature=0.3, json_mode=True)
    if not text:
        return fallback
    try:
        clean = re.sub(r"^```(?:json)?|```$", "", text.strip(),
                       flags=re.I | re.M).strip()
        parsed = json.loads(clean)
        if isinstance(parsed, list) and parsed:
            log.info(" Placement JSON parsed: %d items.", len(parsed))
            return parsed
    except Exception as e:
        log.error(" Placement JSON parse failed: %s | raw: %s", e, text[:300])
    return fallback


def gemini_strategy_advice(meta, matches):
    """Returns (advice_text, ai_was_used_bool)."""
    buckets = meta["buckets"]
    top = ", ".join(m["college"] for m in matches[:3]) or "no strong matches"
    rule = (f"At {meta['effective_percentile']}%ile effective merit "
            f"({meta['category']} category) you have {buckets['Safe']} safe, "
            f"{buckets['Moderate']} moderate and {buckets['Reach']} reach "
            f"options. Lock safe colleges early; keep reach picks to the "
            f"top 1-3 slots.")

    if not GEMINI_KEY:
        return rule, False

    prompt = (
        "You are a blunt MHT-CET CAP-round counsellor.\n"
        f"Student: raw {meta['raw_percentile']}%ile, "
        f"shift-corrected {meta['normalised_percentile']}%ile, "
        f"category {meta['category']}, effective {meta['effective_percentile']}%ile.\n"
        f"Safe/Moderate/Reach: {buckets}. Top matches: {top}.\n"
        "Write 3-4 honest sentences: is the branch expectation realistic, how "
        "to order the CAP form, and whether rising cutoff trends help or hurt "
        "next round. No markdown, no lists, no pleasantries."
    )
    text = _call_gemini(prompt, model="gemini-2.5-flash",
                        max_tokens=512, temperature=0.7)
    return (text or rule), bool(text)


# ==============================================================================
#  ROUTES
# ==============================================================================
@app.route("/", methods=["GET"])
def home():
    return jsonify({
        "service": "Guess My College - Score Engine",
        "version": "2.0",
        "status": "online",
        "supported_exams": list(EXAM_CONFIG.keys()),
        "endpoints": ["/calculate-score", "/predict-college",
                      "/create-payment-order", "/create-pro-order",
                      "/health"],
    })


@app.route("/health", methods=["GET"])
def health():
    return jsonify({
        "status": "ok",
        "ai_advice": "gemini" if _ai_client else "rule-based-fallback",
        "payments": "enabled" if _HAS_CASHFREE else "disabled",
        "database": db_status(),
    })

def _rule_based_percentile(score, max_score, shift=""):
    # FIX E3: NEET sheets without an answer key yield total_score=None — don't crash.
    if score is None:
        return "Score needs the official answer key"
    if max_score != 200:
        return f"~ {(score / max(1, max_score)) * 100:.1f}% Score"
    shift_offset = 2.0 if "s1" in str(shift).lower() else -1.5 if "s2" in str(shift).lower() else 0.0
    eff = max(0, min(200, score + shift_offset))

    if eff >= 160: base = 99.0 + (eff - 160) * (0.99 / 40)
    elif eff >= 140: base = 98.0 + (eff - 140) * (1.0 / 20)
    elif eff >= 120: base = 95.0 + (eff - 120) * (3.0 / 20)
    elif eff >= 100: base = 90.0 + (eff - 100) * (5.0 / 20)
    elif eff >= 80: base = 80.0 + (eff - 80) * (10.0 / 20)
    elif eff >= 50: base = 50.0 + (eff - 50) * (30.0 / 30)
    else: base = eff * (50.0 / 50)

    base = min(99.99, max(1.0, base))
    return f"{max(0.0, base - 0.6):.1f} - {min(99.99, base + 0.4):.1f} %ile"

def predict_percentile_gemini(score, max_score, exam_type, category, shift=""):
    if score is None: return _rule_based_percentile(score, max_score, shift)  # FIX E3
    if not GEMINI_KEY: return _rule_based_percentile(score, max_score, shift)
    shift_ctx = f"Shift: {shift}." if shift else ""
    prompt = f"{category} category scored {score}/{max_score}. {shift_ctx} Predict expected percentile. CRITICAL: Respond ONLY with numbers and '%ile' (e.g. '94.5 - 95.5 %ile'). DO NOT output a single digit."

    text = _call_gemini(prompt, model="gemini-2.5-flash", max_tokens=20, temperature=0.1)
    if text:
        cleaned = text.replace('*', '').replace('`', '').strip()
        if len(cleaned) > 4: return cleaned if "%" in cleaned else cleaned + " %ile"
    return _rule_based_percentile(score, max_score, shift)
def gemini_selftest():
    """Pings Gemini once at startup; prints a clear PASS/FAIL line."""
    if not GEMINI_KEY:
        log.warning(" No Gemini key -- AI features will use the rule-based fallback.")
        return
    log.info(" Running Gemini self-test (gemini-2.5-flash)...")

    # Switched to Flash for reliability
    out = _call_gemini("Reply with exactly one word: PONG", model="gemini-2.5-flash", max_tokens=100, temperature=0)

    if out:
        log.info(" ==> GEMINI IS WORKING. Reply: %s", out)
    else:
        log.error(" ==> GEMINI SELF-TEST FAILED -- read the error lines above.")


@app.route("/calculate-manual", methods=["POST", "OPTIONS"])
def calculate_manual():
    if request.method == "OPTIONS": return ("", 204)
    try:
        data = request.get_json(silent=True) or {}
        marks = float(data.get("marks", 0))
        category = data.get("category", "OPEN").strip()
        shift = data.get("shift", "").strip()
        exam_type = data.get("exam_type", "MHT-CET").strip()

        # Capture advanced splits if provided
        p = float(data.get("p", 0) or 0)
        c = float(data.get("c", 0) or 0)
        m = float(data.get("m", 0) or 0)
        has_splits = bool(p or c or m)

        max_score = 200 if exam_type == "MHT-CET" else 720
        marks = min(max_score, max(0.0, marks))
        predicted_pct = predict_percentile_gemini(marks, max_score, exam_type, category, shift)

        # Look for the calculate_manual route and update the breakdown dictionary
        # Mock the breakdown for the chart if splits exist
        breakdown = {}
        if has_splits:
            breakdown = {
                "Physics": {"score": p, "correct": p, "questions": 50, "max": 50},
                "Chemistry": {"score": c, "correct": c, "questions": 50, "max": 50},
                "Mathematics": {"score": m, "correct": m/2, "questions": 50, "max": 100}
            }

        # Send to Gemini for the massive 3-paragraph response
        result_obj = {"total_score": marks, "max_score": max_score, "breakdown": breakdown, "accuracy": "N/A"}
        advice, ai_used = generate_advice(result_obj, exam_type, category)

        return jsonify({
            "result": "success", "exam_type": exam_type, "category": category,
            "total_score": int(marks) if marks.is_integer() else round(marks, 2),
            "max_score": max_score, "predicted_percentile": predicted_pct,
            "breakdown": breakdown, "has_splits": has_splits,
            "ai_advice": advice, "ai_source": "gemini" if ai_used else "rule-based"
        })
    except Exception as e:
        log.error("Unhandled error in manual calculation:\n%s", traceback.format_exc())
        return error_payload("SERVER_ERROR", "Something went wrong.", str(e), status=500)

@app.route("/calculate-score", methods=["POST", "OPTIONS"])
def calculate_score():
    if request.method == "OPTIONS": return ("", 204)
    try:
        if "file" not in request.files: raise ParseError("NO_FILE", "No file uploaded.")
        file = request.files["file"]
        category = (request.form.get("category") or "OPEN").strip()
        shift = (request.form.get("shift") or "").strip()

        raw = file.read()
        html = extract_html(raw, file.filename)
        soup = make_soup(html)
        exam_type = detect_exam_type(soup) or "MHT-CET"

        questions = parse_mhtcet(soup) if exam_type == "MHT-CET" else parse_neet(soup)[0]
        result = score_questions(questions, exam_type)
        advice, ai_used = generate_advice(result, exam_type, category)
        predicted_pct = predict_percentile_gemini(result["total_score"], result["max_score"], exam_type, category, shift)

        return jsonify({
            "result": "success", "exam_type": exam_type, "category": category,
            **result, "ai_advice": advice, "predicted_percentile": predicted_pct
        })
    except ParseError as e:
        return error_payload(e.code, e.message, e.detail, e.status)
    except Exception as e:
        return error_payload("SERVER_ERROR", "Processing failed.", str(e), status=500)

@app.route("/predict-college", methods=["POST", "OPTIONS"])
def predict_college():
    if request.method == "OPTIONS": return ("", 204)

    # 1. GATEKEEPER: Check Login & Rate Limits first
    client_ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip()
    if not _check_rate_limit(client_ip):
        return jsonify({"result": "error", "error": "Too many requests. Wait a moment."}), 429

    try:
        data = request.get_json(silent=True) or {}
        email = data.get('email')

        # Security: Force Login (now phone-based; email is synthesized phone@phone.gmc)
        if not email:
            return jsonify({"result": "error", "error": "Unauthorized. Please login."}), 401

        # ── ECONOMY GATE (OTP rebuild) ───────────────────────────────────────
        # OLD Google/PRO gate is preserved below (commented) for reference.
        # NEW rule: pro (key-activated) = unlimited. Otherwise consume one of the
        # 4 free predictions. When tokens hit 0 and user isn't pro -> ask for key.
        #
        # --- LEGACY PRO GATE (disabled, kept for reference) ---
        # if not user or not user[0]:                       # is_pro
        #     return jsonify({"result": "error", "error": "PRO Account Required."}), 403
        # if user[1] <= 0:
        #     if user[2] is not None and round(current_score, 2) != round(float(user[2]), 2):
        #         return jsonify({"result": "error", "error": "Tokens empty. Marks locked to previous value."}), 403

        conn = sqlite3.connect(USER_DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        user = cur.execute(
            "SELECT is_pro, tokens_left FROM users WHERE email = ?", (email,)
        ).fetchone()

        if not user:
            conn.close()
            return jsonify({"result": "error", "error": "Account not found. Please log in again."}), 401

        is_pro      = bool(user["is_pro"])
        tokens_left = int(user["tokens_left"] or 0)

        if not is_pro:
            if tokens_left <= 0:
                conn.close()
                return jsonify({
                    "result": "error",
                    "code": "LIMIT_REACHED",
                    "error": "You've used all 4 free predictions. Enter an unlimited key to continue."
                }), 402
            # consume one free prediction
            cur.execute(
                "UPDATE users SET tokens_left = tokens_left - 1 WHERE email = ?", (email,)
            )
            conn.commit()
            tokens_left -= 1
        conn.close()

        # 2. CORE PREDICTION ENGINE
        # legacy support: if only marks supplied, estimate a percentile
        if not data.get("percentile") and not data.get("rank") and data.get("marks"):
            m = float(data.get("marks", 0))
            data["percentile"] = (99.9 if m > 180 else 98.5 if m > 150 else 96.0 if m > 130
                                  else 90.0 if m > 100 else 80.0 if m > 80 else 60.0 + m * 0.2)

        matches, meta = run_prediction(data)

        # ── SAVE STUDENT ROW (OTP rebuild) ───────────────────────────────────
        # One ledger row per prediction: the SuperAdmin export source of truth.
        try:
            _save_student_row(data, meta, matches)
        except Exception as _e:
            log.warning("student row save failed (non-fatal): %s", _e)

        # ---- Sort -----------------------------------------------------------
        sort = (data.get("sort") or "chance").lower()
        if sort == "rank": matches.sort(key=lambda x: -x["cutoff"])
        elif sort == "cutoff": matches.sort(key=lambda x: x["cutoff"])
        else: matches.sort(key=lambda x: (-x["chance"], -x["cutoff"]))

        # ---- AI layer (PARALLEL) --------------------------------------------
        # Both Gemini calls fire at the same time instead of one-after-another,
        # roughly halving the wait. Each has its own safe fallback.
        advice = "Compare multiple categories and branches seamlessly."
        ai_used = False
        placements = []

        want_advice    = data.get("want_advice", True)
        want_placement = data.get("want_placement", True)

        with ThreadPoolExecutor(max_workers=2) as pool:
            fut_advice = pool.submit(gemini_strategy_advice, meta, matches) if want_advice else None
            fut_place  = pool.submit(gemini_placement_summaries, matches[:3]) if want_placement else None

            if fut_advice is not None:
                try:
                    advice, ai_used = fut_advice.result(timeout=45)
                except Exception as e:
                    log.warning("advice future failed: %s", e)

            if fut_place is not None:
                try:
                    placements = fut_place.result(timeout=45)
                except Exception as e:
                    log.warning("placement future failed: %s", e)

        return jsonify({
            "result": "success",
            "meta": meta,
            "matches": matches,
            "placements": placements,
            "ai_advice": advice,
            "ai_source": "gemini" if ai_used else "rule-based",
        })

    except ParseError as e:
        log.warning("Predict ParseError [%s]: %s", e.code, e.detail)
        return error_payload(e.code, e.message, e.detail, e.status)
    except Exception as err:
        log.error("Predict crashed:\n%s", traceback.format_exc())
        return error_payload("PREDICT_FAILED", "Unexpected error.", f"{type(err).__name__}: {err}", status=500)
# --------------------------------------------------------- payment routes -----
@app.route("/create-payment-order", methods=["POST", "OPTIONS"])
def create_payment_order():
    if request.method == "OPTIONS":
        return ("", 204)
    data = request.get_json(silent=True) or {}
    email = data.get("email", "student@example.com")
    return _make_order(amount_rupees=899, email=email, order_note="booking_001")

 # --- Endpoint: Validate Referral Promo Code ---
@app.route("/validate-promo", methods=["POST", "OPTIONS"])
def validate_promo():
    if request.method == "OPTIONS": return ("", 204)
    data = request.get_json(silent=True) or {}
    code = data.get("promo_code", "").strip().upper()
    email = data.get("email", "").strip()

    # FIX I6: a referral code must be tied to a real account.
    if not email:
        return jsonify({"valid": False, "message": "Please log in first."})

    conn = sqlite3.connect(USER_DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # FIX C5: already-PRO users are on the top-up flow where promos don't apply.
    me = cur.execute("SELECT is_pro FROM users WHERE email = ?", (email,)).fetchone()
    if me and me["is_pro"]:
        conn.close()
        return jsonify({"valid": False, "message": "Referral codes don't apply to top-ups."})

    # Check if the promo code belongs to a real user
    referrer = cur.execute("SELECT email FROM users WHERE my_refer_code = ?", (code,)).fetchone()

    if not referrer:
        conn.close()
        return jsonify({"valid": False, "message": "Invalid Referral Code."})

    # Prevent users from using their own code
    if referrer["email"] == email:
        conn.close()
        return jsonify({"valid": False, "message": "You cannot use your own referral code!"})

    # Prevent users from using multiple codes
    user = cur.execute("SELECT referred_by FROM users WHERE email = ?", (email,)).fetchone()
    if user and user["referred_by"]:
        conn.close()
        return jsonify({"valid": False, "message": "You have already used a referral code before."})

    conn.close()
    return jsonify({"valid": True, "discount_pct": 20, "message": "🎉 20% OFF Applied!"})

# --- Endpoint: Dynamic Cashfree Order ---
@app.route('/create-pro-order', methods=['POST', 'OPTIONS'])
def create_pro_order():
    """
    Server-controlled pricing. is_topup is determined by checking the DB,
    not by trusting the client. A user is a top-up if they are already PRO.
    Returns a Cashfree payment_session_id for the frontend SDK.
    """
    if request.method == "OPTIONS": return ("", 204)

    if not _HAS_CASHFREE:
        return error_payload("PAYMENTS_DISABLED",
                             "Online payments are temporarily unavailable.",
                             "Cashfree keys not set.", status=503)

    data = request.get_json(silent=True) or {}
    email      = data.get("email", "").strip()
    promo_code = data.get("promo_code", "").strip().upper()

    # Determine is_topup from DB, not from client-supplied flag
    is_topup = False
    if email:
        conn = sqlite3.connect(USER_DB_PATH)
        conn.row_factory = sqlite3.Row
        user = conn.execute("SELECT is_pro FROM users WHERE email = ?", (email,)).fetchone()
        conn.close()
        is_topup = bool(user and user["is_pro"])

    # Pricing logic (server-controlled), in rupees
    if is_topup:
        amount_rupees = 101            # ₹101 — 40% off, no promo stacking
    elif promo_code:
        conn = sqlite3.connect(USER_DB_PATH)
        referrer = conn.execute(
            "SELECT 1 FROM users WHERE my_refer_code = ?", (promo_code,)
        ).fetchone()
        conn.close()
        amount_rupees = int(169 * 0.80) if referrer else 169
    else:
        amount_rupees = 169            # ₹169 base price

    try:
        session_id, order_id = _cf_create_order(amount_rupees, email, order_note="gmc_pro_order")
        return jsonify({
            "result": "success",
            "payment_session_id": session_id,
            "order_id": order_id,
            "amount": amount_rupees * 100,   # paise, so existing HTML SweetAlerts display right
            "is_topup": is_topup,
            "cashfree_env": CASHFREE_ENV,
        })
    except Exception as e:
        log.error("Cashfree create-pro-order failed: %s", e)
        return jsonify({"result": "error", "error": str(e)}), 502


def _make_order(amount_rupees, email="student@example.com", order_note="gmc_order"):
    if not _HAS_CASHFREE:
        return error_payload(
            "PAYMENTS_DISABLED",
            "Online payments are temporarily unavailable.",
            "Cashfree keys not set.",
            status=503)
    try:
        session_id, order_id = _cf_create_order(amount_rupees, email, order_note=order_note)
        return jsonify({"result": "success",
                        "payment_session_id": session_id,
                        "order_id": order_id,
                        "amount": amount_rupees * 100,   # paise
                        "cashfree_env": CASHFREE_ENV})
    except Exception as e:
        log.error("Cashfree order failed: %s", e)
        return error_payload("PAYMENT_FAILED",
                             "Could not start the payment. Please try again.",
                             str(e), status=502)



# --- Admin: Basic Stats Dashboard (password-protected) ---
# FIX H2: never ship a hardcoded default password. If the env var is unset,
# the endpoint stays disabled rather than falling back to a known string.
ADMIN_PASSWORD = os.environ.get("GMC_ADMIN_PASSWORD")

@app.route("/admin/stats", methods=["GET", "POST", "OPTIONS"])
def admin_stats():
    """
    Admin overview. Password is read from the POST JSON body (preferred) or,
    for backward compatibility, the query string. Set GMC_ADMIN_PASSWORD in env.
    """
    if request.method == "OPTIONS": return ("", 204)

    if not ADMIN_PASSWORD:
        return jsonify({"result": "error", "error": "Admin dashboard is disabled (no password configured)."}), 503

    # FIX H1: accept the password from the POST body so it isn't logged in URLs.
    body = request.get_json(silent=True) or {}
    supplied = body.get("password") or request.args.get("password", "")
    if not hmac.compare_digest(str(supplied), str(ADMIN_PASSWORD)):
        return jsonify({"result": "error", "error": "Unauthorized"}), 401

    conn = sqlite3.connect(USER_DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    total_users    = cur.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    pro_users      = cur.execute("SELECT COUNT(*) FROM users WHERE is_pro = 1").fetchone()[0]
    total_referrals= cur.execute("SELECT SUM(referral_count) FROM users").fetchone()[0] or 0
    new_today      = cur.execute(
        "SELECT COUNT(*) FROM users WHERE DATE(created_at) = DATE('now')"
    ).fetchone()[0]
    recent         = cur.execute(
        "SELECT email, name, is_pro, tokens_left, referral_count, created_at "
        "FROM users ORDER BY created_at DESC LIMIT 20"
    ).fetchall()

    conn.close()
    return jsonify({
        "result": "success",
        "total_users": total_users,
        "pro_users": pro_users,
        "free_users": total_users - pro_users,
        "total_referrals_made": total_referrals,
        "new_signups_today": new_today,
        "recent_users": [dict(r) for r in recent]
    })


# ── NEW (OTP rebuild): SuperAdmin — Student Database ──────────────────────────
def _admin_authorized(req):
    """Shared password check for admin endpoints. Returns (ok, error_response)."""
    if not ADMIN_PASSWORD:
        return False, (jsonify({"result": "error", "error": "Admin disabled (no password configured)."}), 503)
    body = req.get_json(silent=True) or {}
    supplied = body.get("password") or req.args.get("password", "")
    if not hmac.compare_digest(str(supplied), str(ADMIN_PASSWORD)):
        return False, (jsonify({"result": "error", "error": "Unauthorized"}), 401)
    return True, None

@app.route("/admin/students", methods=["GET", "POST", "OPTIONS"])
def admin_students():
    """Returns the full student prediction ledger for the SuperAdmin panel."""
    if request.method == "OPTIONS": return ("", 204)
    ok, err = _admin_authorized(request)
    if not ok: return err

    conn = sqlite3.connect(USER_DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    total      = cur.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    uniq_phone = cur.execute("SELECT COUNT(DISTINCT phone) FROM students").fetchone()[0]
    today      = cur.execute("SELECT COUNT(*) FROM students WHERE DATE(created_at)=DATE('now')").fetchone()[0]
    rows = cur.execute(
        "SELECT id, name, phone, percentile, category, branch, city, shift, "
        "cap_round, total_matches, created_at FROM students ORDER BY id DESC LIMIT 1000"
    ).fetchall()
    conn.close()

    return jsonify({
        "result": "success",
        "total_predictions": total,
        "unique_students": uniq_phone,
        "predictions_today": today,
        "students": [dict(r) for r in rows],
    })

@app.route("/admin/export-students", methods=["POST", "OPTIONS"])
def admin_export_students():
    """Streams the full student ledger as a real .xlsx file."""
    if request.method == "OPTIONS": return ("", 204)
    ok, err = _admin_authorized(request)
    if not ok: return err

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
    except Exception as e:
        return jsonify({"result": "error", "error": f"openpyxl not installed: {e}"}), 500

    conn = sqlite3.connect(USER_DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT id, name, phone, percentile, category, branch, city, shift, "
        "cap_round, total_matches, created_at FROM students ORDER BY id DESC"
    ).fetchall()
    conn.close()

    wb = Workbook(); ws = wb.active; ws.title = "Students"
    headers = ["ID", "Name", "Phone", "Percentile", "Category", "Branch",
               "City", "Shift", "CAP Round", "Matches", "Date/Time"]
    ws.append(headers)
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="6D28D9")
    for c in ws[1]:
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r["id"], r["name"], r["phone"], r["percentile"], r["category"],
                   r["branch"], r["city"], r["shift"], r["cap_round"],
                   r["total_matches"], r["created_at"]])
    widths = [6, 22, 16, 12, 14, 12, 16, 12, 10, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from flask import Response
    ts = time.strftime("%Y%m%d_%H%M")
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=GMC_Students_{ts}.xlsx"}
    )


@app.errorhandler(413)
def too_large(_):
    return error_payload("FILE_TOO_LARGE",
                         "That file is too large (limit is 64 MB).",
                         "MAX_CONTENT_LENGTH exceeded.", status=413)


@app.errorhandler(404)
def not_found(_):
    return error_payload("NOT_FOUND", "That endpoint does not exist.",
                         "404", status=404)

@app.route('/cashfree-webhook', methods=['POST'])
def cashfree_webhook():
    try:
        # 1. Grab the security headers Cashfree sends
        timestamp = request.headers.get('x-webhook-timestamp')
        actual_signature = request.headers.get('x-webhook-signature')

        # 2. Get the raw text of the message
        raw_body = request.data.decode('utf-8')

        # 3. VERIFY IT'S ACTUALLY CASHFREE (signature check)
        if not CASHFREE_SECRET_KEY:
            log.error("Webhook received but CASHFREE_SECRET_KEY is not set.")
            return jsonify({"error": "Server not configured"}), 500

        signed_payload = (timestamp or "") + raw_body
        computed_signature = base64.b64encode(
            hmac.new(CASHFREE_SECRET_KEY.encode('utf-8'),
                     signed_payload.encode('utf-8'),
                     hashlib.sha256).digest()
        ).decode('utf-8')

        if not actual_signature or not hmac.compare_digest(computed_signature, actual_signature):
            log.warning("ALERT: Fake webhook attempt blocked!")
            return jsonify({"error": "Invalid signature"}), 400

        # 4. Signature valid → read the data
        webhook_data = request.json or {}

        # Only act on a successful payment event
        if webhook_data.get('type') == 'PAYMENT_SUCCESS_WEBHOOK':
            order = webhook_data.get('data', {}).get('order', {})
            customer = webhook_data.get('data', {}).get('customer_details', {})
            order_id = order.get('order_id')
            customer_email = customer.get('customer_email')
            order_note = order.get('order_note', '')

            log.info("WEBHOOK SUCCESS  order=%s  email=%s  note=%s",
                     order_id, customer_email, order_note)

            # Only PRO orders should grant PRO. One-off counselling/booking
            # payments (booking_001) must NOT upgrade the user to PRO.
            if customer_email and order_note != "booking_001":
                # Mirror /verify-and-grant: top-up if already PRO, else fresh grant.
                conn = sqlite3.connect(USER_DB_PATH)
                conn.row_factory = sqlite3.Row
                user = conn.execute(
                    "SELECT is_pro FROM users WHERE email = ?", (customer_email,)
                ).fetchone()
                conn.close()
                is_topup = bool(user and user["is_pro"])

                # promo_code isn't available on the webhook path, so referral
                # rewards only fire via /verify-and-grant. PRO + tokens still granted.
                _do_grant_pro(customer_email, promo_code="", is_topup=is_topup, order_id=order_id)
                log.info("✅  PRO granted via webhook  email=%s  topup=%s",
                         customer_email, is_topup)

        # Always return 200 so Cashfree stops retrying
        return jsonify({"status": "received"}), 200

    except Exception as e:
        log.error("Webhook Error: %s", e)
        return jsonify({"error": "Webhook failed"}), 500

# --- Endpoint: Redeem Custom Product Key ---
@app.route("/redeem-key", methods=["POST", "OPTIONS"])
def redeem_key():
    if request.method == "OPTIONS": return ("", 204)
    data = request.get_json(silent=True) or {}
    
    email = data.get("email", "").strip()
    key = data.get("key", "").strip().upper()

    if not email or not key:
        return jsonify({"result": "error", "error": "Email and Product Key are required."}), 400

    # FIX I2: keep master keys out of source. Set GMC_PRODUCT_KEYS in the env as a
    # comma-separated list (e.g. "BRO-PRO-MAX,FRIEND-PASS-2026"). Falls back to the
    # legacy hardcoded set only if the env var is missing, so nothing breaks on deploy.
    _env_keys = os.environ.get("GMC_PRODUCT_KEYS", "")
    if _env_keys.strip():
        VALID_KEYS = [k.strip().upper() for k in _env_keys.split(",") if k.strip()]
    else:
        VALID_KEYS = ["BRO-PRO-MAX", "FRIEND-PASS-2026", "PARTH-VIP"]

    if key not in VALID_KEYS:
        return jsonify({"result": "error", "error": "Invalid or expired Product Key."}), 403

    try:
        # Connect to the User Database
        conn = sqlite3.connect(USER_DB_PATH)
        cur = conn.cursor()

        # Grant PRO, give them 99999 tokens (Unlimited), and wipe any locked marks
        cur.execute(
            "UPDATE users SET is_pro = 1, tokens_left = 99999, locked_marks = NULL WHERE email = ?",
            (email,)
        )

        # FIX I1: if no row changed, the account doesn't exist yet — don't fake success.
        if cur.rowcount == 0:
            conn.close()
            return jsonify({"result": "error", "error": "Account not found. Please log in with Google first, then redeem."}), 404

        conn.commit()
        conn.close()

        return jsonify({"result": "success", "message": "VIP PRO Access Unlocked!"})
        
    except Exception as e:
        log.error(f"Redeem Key Error for {email}: {e}")
        return jsonify({"result": "error", "error": "Server error while redeeming key."}), 500

if __name__ == "__main__":
    log.info("Score Engine v2 starting on http://127.0.0.1:5000")
    gemini_selftest()          # prints PASS/FAIL for Gemini at startup
    # debug=False for production. Set to True only for local development.
    app.run(host="127.0.0.1", port=5000, debug=False)